import openpyxl, os

DESKTOP = 'c:/Users/22430/Desktop/高考河北数据'

for fname in os.listdir(DESKTOP):
    if '2025' in fname and fname.endswith('.xlsx'):
        path = os.path.join(DESKTOP, fname)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"\n{'='*60}")
            print(f"FILE: {fname} [{sname}] {ws.max_row}行 × {ws.max_column}列")
            for col in range(1, min(ws.max_column+1, 16)):
                v = ws.cell(row=1, column=col).value
                if v: print(f"  Col{col}: {str(v)[:60]}")
            for r in [2, 3, ws.max_row//2, ws.max_row]:
                if r > ws.max_row: continue
                vals = []
                for col in range(1, min(ws.max_column+1, 12)):
                    v = ws.cell(row=r, column=col).value
                    if v: vals.append(str(v)[:40])
                print(f"  Row{r}: {vals}")
        wb.close()

if '25年' in ''.join(os.listdir(DESKTOP)):
    fname = [f for f in os.listdir(DESKTOP) if '25年' in f][0]
    path = os.path.join(DESKTOP, fname)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n{'='*60}")
        print(f"FILE: {fname} [{sname}] {ws.max_row}行 × {ws.max_column}列")
        for col in range(1, min(ws.max_column+1, 16)):
            v = ws.cell(row=1, column=col).value
            if v: print(f"  Col{col}: {str(v)[:60]}")
        for r in [2, 3, ws.max_row//2, ws.max_row]:
            if r > ws.max_row: continue
            vals = []
            for col in range(1, min(ws.max_column+1, 12)):
                v = ws.cell(row=r, column=col).value
                if v: vals.append(str(v)[:40])
            print(f"  Row{r}: {vals}")
    wb.close()
