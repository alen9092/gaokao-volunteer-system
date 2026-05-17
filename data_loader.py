"""
One-time data import: Excel -> SQLite.
Run: python -u data_loader.py
"""
import os, sys
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from app import create_app
from models import db, School, AdmissionRecord, ScoreDistribution, EnrollmentPlan

app = create_app()
DESKTOP = 'c:/Users/22430/Desktop/高考河北数据'
D_DRIVE = 'd:/高考数据'


def to_int(v):
    if v is None: return None
    try: return int(float(str(v)))
    except: return None


def to_bool(v):
    return str(v).strip() in ('是', '1', 'True', 'true')


def ns(v):
    """Normalize string."""
    return str(v).strip() if v else ''


def normalize_subject(raw):
    s = str(raw).strip()
    if '物理' in s: return '物理类'
    if '历史' in s: return '历史类'
    return s


def file_matching(pattern):
    """Find a file in desktop whose name contains pattern."""
    for f in os.listdir(DESKTOP):
        if pattern in f and f.endswith('.xlsx'):
            return os.path.join(DESKTOP, f)
    return None


# ===================== STEP 1: BUILD SCHOOL DATABASE =====================

def build_schools():
    """Extract unique schools with metadata from 2024 undergrad data + 2025 major data."""
    schools = {}  # code -> dict

    # Source A: 2024 undergrad data (has 985/211/双一流 metadata but no school codes)
    for fname in os.listdir(DESKTOP):
        if '专业录取数据-本科' not in fname: continue
        path = os.path.join(DESKTOP, fname)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Headers on row 2
        cols = {}
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(row=2, column=col).value or '').strip()
            for key, label in [
                ('name', '学校'), ('type', '学校类型'), ('attr', '学校属性'),
                ('level', '办学层次'), ('city', '城市'), ('province', '省份'),
                ('985', '985'), ('211', '211'), ('double', '双一流')]:
                if h == label: cols[key] = col

        if 'name' not in cols:
            wb.close()
            continue

        for r in range(3, ws.max_row + 1):
            name = ns(ws.cell(row=r, column=cols['name']).value)
            if not name: continue
            # Use name as temporary key; codes come later
            key = name
            if key in schools: continue

            schools[key] = {
                'code': '', 'name': name,
                'type': ns(ws.cell(row=r, column=cols.get('type', 0)).value) if 'type' in cols else '',
                'attribute': ns(ws.cell(row=r, column=cols.get('attr', 0)).value) if 'attr' in cols else '',
                'level': ns(ws.cell(row=r, column=cols.get('level', 0)).value) if 'level' in cols else '',
                'city': ns(ws.cell(row=r, column=cols.get('city', 0)).value) if 'city' in cols else '',
                'province': ns(ws.cell(row=r, column=cols.get('province', 0)).value) if 'province' in cols else '',
                'is_985': to_bool(ws.cell(row=r, column=cols.get('985', 0)).value) if '985' in cols else False,
                'is_211': to_bool(ws.cell(row=r, column=cols.get('211', 0)).value) if '211' in cols else False,
                'is_double_first': to_bool(ws.cell(row=r, column=cols.get('double', 0)).value) if 'double' in cols else False,
            }
        wb.close()

    # Source B: 2025 major data (has 院校代码 + 院校名称)
    f_2025 = file_matching('25年全国高校')
    if f_2025:
        wb = openpyxl.load_workbook(f_2025, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Headers on row 1
        cols = {}
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(row=1, column=col).value or '')
            if '院校代码' in h: cols['code'] = col
            if '院校名称' in h: cols['name'] = col
            if '所在省' in h or ('省份' in h and 'col' not in locals()): cols['province'] = col
            if '公私' in h or '性质' in h: cols['attr'] = col
            if '层次' in h: cols['level'] = col

        for r in range(2, ws.max_row + 1):
            code = ns(ws.cell(row=r, column=cols.get('code', 0)).value) if 'code' in cols else ''
            name = ns(ws.cell(row=r, column=cols.get('name', 0)).value) if 'name' in cols else ''
            if not name: continue

            if name in schools:
                if code and not schools[name]['code']:
                    schools[name]['code'] = code
                if 'attr' in cols and not schools[name]['attribute']:
                    schools[name]['attribute'] = ns(ws.cell(row=r, column=cols['attr']).value)
                if 'province' in cols and not schools[name]['province']:
                    schools[name]['province'] = ns(ws.cell(row=r, column=cols['province']).value)
            else:
                schools[name] = {
                    'code': code, 'name': name,
                    'type': '', 'city': '',
                    'attribute': ns(ws.cell(row=r, column=cols.get('attr', 0)).value) if 'attr' in cols else '',
                    'level': ns(ws.cell(row=r, column=cols.get('level', 0)).value) if 'level' in cols else '',
                    'province': ns(ws.cell(row=r, column=cols.get('province', 0)).value) if 'province' in cols else '',
                    'is_985': False, 'is_211': False, 'is_double_first': False,
                }
        wb.close()

    # Source C: 2023 school-level file (has code mapping)
    f_2023 = file_matching('2023-河北-院校')
    if f_2023:
        wb = openpyxl.load_workbook(f_2023, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        cols = {}
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(row=1, column=col).value or '')
            if '院校代码' in h: cols['code'] = col
            if '院校名称' in h: cols['name'] = col
        for r in range(2, ws.max_row + 1):
            code = ns(ws.cell(row=r, column=cols.get('code', 0)).value) if 'code' in cols else ''
            name = ns(ws.cell(row=r, column=cols.get('name', 0)).value) if 'name' in cols else ''
            if name and name in schools and code and not schools[name]['code']:
                schools[name]['code'] = code
        wb.close()

    # Build name->code map, generate codes for missing ones
    name_to_code = {}
    for key, s in schools.items():
        if s['code']:
            name_to_code[s['name']] = s['code']
    for key, s in schools.items():
        if not s['code']:
            code = f'X{abs(hash(s["name"])) % 9000 + 1000:04d}'
            s['code'] = code
        name_to_code[s['name']] = s['code']

    # Deduplicate by code
    by_code = {}
    for s in schools.values():
        c = s['code']
        if c not in by_code or s['is_985'] or s['is_211']:
            by_code[c] = s

    print(f"  {len(by_code)} unique schools (from {len(schools)} names)")
    return list(by_code.values()), name_to_code


# ===================== STEP 2: LOAD ADMISSION RECORDS =====================

def load_2023_admissions(name_to_code):
    records = []
    f = file_matching('2023河北专业')
    if not f: return records
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cols = {}
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=col).value or '')
        for key, label in [('year', '年份'), ('subject', '科类'), ('batch', '批次'),
                           ('school', '院校名称'), ('major', '专业名称'),
                           ('score', '最低分'), ('rank', '最低位次'), ('note', '专业备注')]:
            if label in h and key not in cols: cols[key] = col

    for r in range(2, ws.max_row + 1):
        score = to_int(ws.cell(row=r, column=cols.get('score', 0)).value) if 'score' in cols else None
        if not score or score < 100: continue
        name = ns(ws.cell(row=r, column=cols.get('school', 0)).value) if 'school' in cols else ''
        records.append({
            'year': to_int(ws.cell(row=r, column=cols.get('year', 0)).value) or 2023,
            'subject': normalize_subject(ns(ws.cell(row=r, column=cols.get('subject', 0)).value)) if 'subject' in cols else '',
            'batch': ns(ws.cell(row=r, column=cols.get('batch', 0)).value) if 'batch' in cols else '',
            'school_code': name_to_code.get(name, ''),
            'school_name': name,
            'major_code': '',
            'major_name': ns(ws.cell(row=r, column=cols.get('major', 0)).value) if 'major' in cols else '',
            'min_score': score,
            'min_rank': to_int(ws.cell(row=r, column=cols.get('rank', 0)).value) if 'rank' in cols else None,
            'subject_req': ns(ws.cell(row=r, column=cols.get('note', 0)).value)[:50] if 'note' in cols else '',
        })
    wb.close()
    print(f"  2023 major: {len(records)} records")
    return records


def load_2024_admissions(name_to_code, track):
    """track = '物理类' or '历史类'"""
    records = []
    for fname in os.listdir(DESKTOP):
        if '专业录取数据' not in fname or '2024' not in fname: continue
        if track.replace('类', '') not in fname: continue
        path = os.path.join(DESKTOP, fname)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Headers on row 2
        cols = {}
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(row=2, column=col).value or '')
            for key, label in [('school', '学校'), ('major', '专业'), ('score', '最低分'),
                               ('rank', '最低位次'), ('batch', '录取批次'), ('req', '选科')]:
                if h == label: cols[key] = col

        batch_default = '本科批' if '本科' in fname else '专科批'
        for r in range(3, ws.max_row + 1):
            score = to_int(ws.cell(row=r, column=cols.get('score', 0)).value) if 'score' in cols else None
            if not score or score < 100: continue
            name = ns(ws.cell(row=r, column=cols.get('school', 0)).value) if 'school' in cols else ''
            records.append({
                'year': 2024, 'subject': track,
                'batch': ns(ws.cell(row=r, column=cols.get('batch', 0)).value) if 'batch' in cols else batch_default,
                'school_code': name_to_code.get(name, ''),
                'school_name': name,
                'major_code': '',
                'major_name': ns(ws.cell(row=r, column=cols.get('major', 0)).value) if 'major' in cols else '',
                'min_score': score,
                'min_rank': to_int(ws.cell(row=r, column=cols.get('rank', 0)).value) if 'rank' in cols else None,
                'subject_req': ns(ws.cell(row=r, column=cols.get('req', 0)).value)[:50] if 'req' in cols else '',
            })
        wb.close()
    print(f"  2024 {track}: {len(records)} records")
    return records


def load_2025_major(name_to_code):
    """Load 2025 major data from 25年全国高校在河北省专业录取数据.xlsx"""
    records = []
    f = file_matching('25年全国高校')
    if not f: return records
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cols = {}
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=col).value or '')
        for key, label in [('school', '院校名称'), ('scode', '院校代码'), ('major', '专业'),
                           ('mcode', '专业代码'), ('score', '最低分'), ('rank', '最低位次'),
                           ('batch', '层次'), ('subject', '科类'), ('req', '选科要求'),
                           ('note', '专业备注')]:
            if label in h and key not in cols: cols[key] = col

    for r in range(2, ws.max_row + 1):
        score = to_int(ws.cell(row=r, column=cols.get('score', 0)).value) if 'score' in cols else None
        if not score or score < 100: continue
        name = ns(ws.cell(row=r, column=cols.get('school', 0)).value) if 'school' in cols else ''
        batch_raw = ns(ws.cell(row=r, column=cols.get('batch', 0)).value) if 'batch' in cols else ''
        # Normalize batch
        if '本科' in batch_raw:
            batch = '本科批'
        elif '专科' in batch_raw or '高职' in batch_raw:
            batch = '专科批'
        else:
            batch = batch_raw
        records.append({
            'year': 2025, 'subject': normalize_subject(ns(ws.cell(row=r, column=cols.get('subject', 0)).value)),
            'batch': batch,
            'school_code': ns(ws.cell(row=r, column=cols.get('scode', 0)).value) if 'scode' in cols else '',
            'school_name': name,
            'major_code': ns(ws.cell(row=r, column=cols.get('mcode', 0)).value) if 'mcode' in cols else '',
            'major_name': ns(ws.cell(row=r, column=cols.get('major', 0)).value) if 'major' in cols else '',
            'min_score': score,
            'min_rank': to_int(ws.cell(row=r, column=cols.get('rank', 0)).value) if 'rank' in cols else None,
            'subject_req': ns(ws.cell(row=r, column=cols.get('req', 0)).value)[:50] if 'req' in cols else '',
        })
    wb.close()
    print(f"  2025 major: {len(records)} records")
    return records


def load_2025_vocational(name_to_code):
    """Official Hebei 2025 vocational data from D: drive."""
    records = []
    for fname in os.listdir(D_DRIVE):
        if '2025' not in fname or not fname.endswith('.xlsx'): continue
        if '物理' not in fname and '历史' not in fname: continue
        path = os.path.join(D_DRIVE, fname)
        track = '物理类' if '物理' in fname else '历史类'
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        cols = {}
        for col in range(1, min(ws.max_column + 1, 20)):
            h = str(ws.cell(row=1, column=col).value or '')
            if '院校代码' in h: cols['scode'] = col
            if '院校名称' in h: cols['school'] = col
            if '专业代码' in h: cols['mcode'] = col
            if '专业名称' in h: cols['major'] = col
            if '最低分' in h or '投档' in h: cols['score'] = col

        # 2025 data doesn't have rank - we'll cross-reference later
        for r in range(2, ws.max_row + 1):
            score = to_int(ws.cell(row=r, column=cols.get('score', 0)).value) if 'score' in cols else None
            if not score or score < 100: continue
            records.append({
                'year': 2025, 'subject': track, 'batch': '专科批',
                'school_code': ns(ws.cell(row=r, column=cols.get('scode', 0)).value) if 'scode' in cols else '',
                'school_name': ns(ws.cell(row=r, column=cols.get('school', 0)).value) if 'school' in cols else '',
                'major_code': ns(ws.cell(row=r, column=cols.get('mcode', 0)).value) if 'mcode' in cols else '',
                'major_name': ns(ws.cell(row=r, column=cols.get('major', 0)).value) if 'major' in cols else '',
                'min_score': score,
                'min_rank': None,  # 2025 vocational doesn't have rank in source
                'subject_req': '',
            })
        wb.close()
    # Deduplicate - 2025 major data might already cover some of these
    print(f"  2025 vocational (official): {len(records)} records (may overlap with 2025 major)")
    return records


# ===================== STEP 3: SCORE DISTRIBUTION =====================

def load_score_distribution():
    records = []
    path = os.path.join(DESKTOP, [f for f in os.listdir(DESKTOP) if '一分一段' in f and f.endswith('.xlsx')][0])

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        if '对比' in sname or 'vs' in sname.lower(): continue
        year = 2024 if '2024' in sname else 2025
        subject = '物理类' if '物理' in sname else '历史类'

        ws = wb[sname]
        # Find the data: score in one column, cumulative at score+2 or score+3
        for r in range(2, min(ws.max_row + 1, 550)):
            # Try to find numeric values
            score = count = cum = None
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if isinstance(v, (int, float)) and 100 <= v <= 750:
                    score = int(v)
                    cnt_v = ws.cell(row=r, column=col + 1).value
                    cum_v = ws.cell(row=r, column=col + 2).value
                    if isinstance(cnt_v, (int, float)): count = int(cnt_v)
                    if isinstance(cum_v, (int, float)): cum = int(cum_v)
                    break
            if score and cum:
                records.append({
                    'year': year, 'subject': subject,
                    'score': score, 'count': count or 0, 'cumulative': cum,
                })

    wb.close()

    # Also load 2023 from D: drive file
    path_2023 = os.path.join(D_DRIVE, '2023-2025河北高考投档线大全.xlsx')
    if os.path.exists(path_2023):
        wb = openpyxl.load_workbook(path_2023, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            if '2023一分一段' not in sname: continue
            ws = wb[sname]
            subject = '物理类' if '物理' in sname else '历史类'
            for r in range(4, ws.max_row + 1):
                score = to_int(ws.cell(row=r, column=1).value)
                cnt = to_int(ws.cell(row=r, column=3).value)
                cum = to_int(ws.cell(row=r, column=4).value)
                if score and cum:
                    records.append({'year': 2023, 'subject': subject,
                                    'score': score, 'count': cnt or 0, 'cumulative': cum})
        wb.close()

    print(f"  Score distribution: {len(records)} records")
    return records


# ===================== STEP 4: ENROLLMENT PLANS =====================

def load_enrollment_plans():
    records = []

    # 2024 from desktop
    f_2024 = file_matching('河北省2024年高考招生计划')
    if f_2024:
        wb = openpyxl.load_workbook(f_2024, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            if '总表' in sname: continue
            ws = wb[sname]
            # Row 1 is headers (年份, 录取批次, 科类名称, 计划性质, ...)
            cols = {}
            for col in range(1, ws.max_column + 1):
                h = str(ws.cell(row=1, column=col).value or '')
                if '院校代码' in h: cols['scode'] = col
                if '院校名称' in h: cols['school'] = col
                if '专业代码' in h: cols['mcode'] = col
                if '专业名称' in h: cols['major'] = col
                if '计划数' in h: cols['plan'] = col
                if '选科' in h: cols['req'] = col
                if '学费' in h: cols['tuition'] = col
                if '学制' in h: cols['duration'] = col

            for r in range(2, ws.max_row + 1):
                sn = ns(ws.cell(row=r, column=cols.get('school', 0)).value) if 'school' in cols else ''
                mn = ns(ws.cell(row=r, column=cols.get('major', 0)).value) if 'major' in cols else ''
                if not sn and not mn: continue
                records.append({
                    'year': 2024, 'batch': sname, 'subject': '', 'plan_type': '',
                    'school_code': ns(ws.cell(row=r, column=cols.get('scode', 0)).value) if 'scode' in cols else '',
                    'school_name': sn,
                    'major_code': ns(ws.cell(row=r, column=cols.get('mcode', 0)).value) if 'mcode' in cols else '',
                    'major_name': mn,
                    'plan_count': to_int(ws.cell(row=r, column=cols.get('plan', 0)).value) or 0 if 'plan' in cols else 0,
                    'subject_req': ns(ws.cell(row=r, column=cols.get('req', 0)).value)[:50] if 'req' in cols else '',
                    'tuition': to_int(ws.cell(row=r, column=cols.get('tuition', 0)).value) if 'tuition' in cols else None,
                    'duration': to_int(ws.cell(row=r, column=cols.get('duration', 0)).value) if 'duration' in cols else None,
                })
        wb.close()

    # 2025 from desktop
    f_2025 = file_matching('2025年河北省招生计划')
    if f_2025:
        wb = openpyxl.load_workbook(f_2025, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Row 2 is headers
        cols = {}
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(row=2, column=col).value or '')
            if '院校代码' in h: cols['scode'] = col
            if '院校名称' in h: cols['school'] = col
            if '专业代码' in h: cols['mcode'] = col
            if '专业名称' in h: cols['major'] = col
            if '计划数' in h: cols['plan'] = col
            if '选科' in h: cols['req'] = col
            if '批次' in h: cols['batch'] = col
            if '科目' in h: cols['subject'] = col
            if '学费' in h: cols['tuition'] = col
            if '学制' in h: cols['duration'] = col

        for r in range(3, ws.max_row + 1):
            sn = ns(ws.cell(row=r, column=cols.get('school', 0)).value) if 'school' in cols else ''
            mn = ns(ws.cell(row=r, column=cols.get('major', 0)).value) if 'major' in cols else ''
            if not sn and not mn: continue
            records.append({
                'year': 2025, 'subject': ns(ws.cell(row=r, column=cols.get('subject', 0)).value) if 'subject' in cols else '',
                'batch': ns(ws.cell(row=r, column=cols.get('batch', 0)).value) if 'batch' in cols else '',
                'plan_type': '',
                'school_code': ns(ws.cell(row=r, column=cols.get('scode', 0)).value) if 'scode' in cols else '',
                'school_name': sn,
                'major_code': ns(ws.cell(row=r, column=cols.get('mcode', 0)).value) if 'mcode' in cols else '',
                'major_name': mn,
                'plan_count': to_int(ws.cell(row=r, column=cols.get('plan', 0)).value) or 0 if 'plan' in cols else 0,
                'subject_req': ns(ws.cell(row=r, column=cols.get('req', 0)).value)[:50] if 'req' in cols else '',
                'tuition': to_int(ws.cell(row=r, column=cols.get('tuition', 0)).value) if 'tuition' in cols else None,
                'duration': to_int(ws.cell(row=r, column=cols.get('duration', 0)).value) if 'duration' in cols else None,
            })
        wb.close()

    # 2023 from desktop
    f_2023 = file_matching('河北-2023-招生计划')
    if f_2023:
        wb = openpyxl.load_workbook(f_2023, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        cols = {}
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(row=1, column=col).value or '')
            if '院校代码' in h: cols['scode'] = col
            if '院校名称' in h: cols['school'] = col
            if '专业代码' in h: cols['mcode'] = col
            if '专业名称' in h: cols['major'] = col
            if '计划数' in h: cols['plan'] = col
            if '选科' in h: cols['req'] = col
            if '批次' in h: cols['batch'] = col
            if '科类' in h: cols['subject'] = col
            if '学费' in h: cols['tuition'] = col
            if '学制' in h: cols['duration'] = col

        for r in range(2, ws.max_row + 1):
            sn = ns(ws.cell(row=r, column=cols.get('school', 0)).value) if 'school' in cols else ''
            mn = ns(ws.cell(row=r, column=cols.get('major', 0)).value) if 'major' in cols else ''
            if not sn and not mn: continue
            records.append({
                'year': 2023, 'subject': ns(ws.cell(row=r, column=cols.get('subject', 0)).value) if 'subject' in cols else '',
                'batch': ns(ws.cell(row=r, column=cols.get('batch', 0)).value) if 'batch' in cols else '',
                'plan_type': '',
                'school_code': ns(ws.cell(row=r, column=cols.get('scode', 0)).value) if 'scode' in cols else '',
                'school_name': sn,
                'major_code': ns(ws.cell(row=r, column=cols.get('mcode', 0)).value) if 'mcode' in cols else '',
                'major_name': mn,
                'plan_count': to_int(ws.cell(row=r, column=cols.get('plan', 0)).value) or 0 if 'plan' in cols else 0,
                'subject_req': ns(ws.cell(row=r, column=cols.get('req', 0)).value)[:50] if 'req' in cols else '',
                'tuition': to_int(ws.cell(row=r, column=cols.get('tuition', 0)).value) if 'tuition' in cols else None,
                'duration': to_int(ws.cell(row=r, column=cols.get('duration', 0)).value) if 'duration' in cols else None,
            })
        wb.close()

    print(f"  Enrollment plans: {len(records)} records")
    return records


# ===================== MAIN =====================

def main():
    print("=" * 60)
    print("Gaokao Data Import")
    print("=" * 60)

    with app.app_context():
        # Clear
        print("\n[1] Clearing existing data...")
        for t in [AdmissionRecord, ScoreDistribution, EnrollmentPlan, School]:
            db.session.execute(db.delete(t))
        db.session.commit()

        # Schools
        print("\n[2] Building school database...")
        school_list, name_to_code = build_schools()
        db.session.bulk_save_objects([School(**s) for s in school_list])
        db.session.commit()

        # Admission records
        all_adm = []
        print("\n[3] Loading 2023 admissions...")
        all_adm += load_2023_admissions(name_to_code)

        print("\n[4] Loading 2024 admissions...")
        all_adm += load_2024_admissions(name_to_code, '物理类')
        all_adm += load_2024_admissions(name_to_code, '历史类')

        print("\n[5] Loading 2025 admissions...")
        all_adm += load_2025_major(name_to_code)  # 48K records!
        print("\n[6] Loading 2025 vocational (official)...")
        all_adm += load_2025_vocational(name_to_code)

        # Remove duplicates (same year+subject+school+score+major)
        seen = set()
        deduped = []
        for r in all_adm:
            key = (r['year'], r['subject'], r['min_score'], r['school_name'][:20], r['major_name'][:20])
            if key not in seen:
                seen.add(key)
                deduped.append(r)
        print(f"\n[7] Total: {len(all_adm)} raw, {len(deduped)} deduped")

        print(f"\n[8] Inserting {len(deduped)} admission records...")
        batch_size = 2000
        for i in range(0, len(deduped), batch_size):
            batch = deduped[i:i + batch_size]
            db.session.bulk_insert_mappings(AdmissionRecord, batch)
            if i % 10000 == 0:
                db.session.commit()
                print(f"  ... {i}/{len(deduped)}")
        db.session.commit()

        # Score distribution
        print("\n[9] Loading score distribution...")
        sd = load_score_distribution()
        db.session.bulk_insert_mappings(ScoreDistribution, sd)
        db.session.commit()

        # Enrollment plans
        print("\n[10] Loading enrollment plans...")
        ep = load_enrollment_plans()
        for i in range(0, len(ep), batch_size):
            db.session.bulk_insert_mappings(EnrollmentPlan, ep[i:i + batch_size])
            if i % 20000 == 0:
                db.session.commit()
        db.session.commit()

        # Summary
        print("\n" + "=" * 60)
        print("IMPORT COMPLETE")
        print(f"  Schools:            {School.query.count()}")
        print(f"  Admission Records:  {AdmissionRecord.query.count()}")
        print(f"  Score Distribution: {ScoreDistribution.query.count()}")
        print(f"  Enrollment Plans:   {EnrollmentPlan.query.count()}")
        print("=" * 60)


if __name__ == '__main__':
    main()
