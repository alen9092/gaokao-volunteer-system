"""
Simplified data import - one file at a time with progress.
Run: python -u import_data.py
"""
import os, sys, time
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from app import create_app
from models import db, School, AdmissionRecord, ScoreDistribution, EnrollmentPlan

app = create_app()
DSK = 'c:/Users/22430/Desktop/高考河北数据'
DDRV = 'd:/高考数据'


def ni(v):
    """int or None"""
    if v is None: return None
    try: return int(float(str(v)))
    except: return None


def ns(v):
    return str(v).strip() if v else ''


def find_header(ws):
    """Return (header_row, {col_name: col_index})."""
    for r in (1, 2):
        c1 = ns(ws.cell(row=r, column=1).value)
        if ('学校' in c1 or '院校' in c1 or '年份' in c1 or '定向' in c1):
            cols = {}
            for col in range(1, min(ws.max_column + 1, 20)):
                h = ns(ws.cell(row=r, column=col).value)
                if h:
                    cols[h] = col
            return r, cols
    return 1, {}


def file_by_kw(*keywords):
    for f in os.listdir(DSK):
        if all(k in f for k in keywords) and f.endswith('.xlsx'):
            return os.path.join(DSK, f)
    return None


# ===================== MAIN =====================

def main():
    with app.app_context():
        # Clear
        print("[CLEAR]")
        for t in (AdmissionRecord, ScoreDistribution, EnrollmentPlan, School):
            db.session.execute(db.delete(t))
        db.session.commit()

        # ---- SCHOOLS ----
        print("\n[SCHOOLS] Extracting from 2024 undergrad + 2025 major data...")
        schools_by_code = {}

        # Source 1: 2024 undergrad (has 985/211 tags, headers on row 1)
        for fname in os.listdir(DSK):
            if '专业录取数据-本科' not in fname: continue
            path = os.path.join(DSK, fname)
            print(f"  Reading {fname}...")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hr, cols = find_header(ws)

            name_col = cols.get('学校', None)
            if not name_col: name_col = cols.get('院校名称', None)
            if not name_col:
                wb.close()
                continue

            for r in range(hr + 1, ws.max_row + 1):
                name = ns(ws.cell(row=r, column=name_col).value)
                if not name: continue
                code = schools_by_code.get('_name_' + name, {}).get('code', '')
                if not code:
                    code = f'X{abs(hash(name)) % 9000 + 1000:04d}'

                key = code
                if key in schools_by_code: continue

                schools_by_code[key] = {
                    'code': code, 'name': name,
                    'type': ns(ws.cell(row=r, column=cols.get('学校类型', 0)).value) if '学校类型' in cols else '',
                    'attribute': ns(ws.cell(row=r, column=cols.get('学校属性', 0)).value) if '学校属性' in cols else '',
                    'level': ns(ws.cell(row=r, column=cols.get('办学层次', 0)).value) if '办学层次' in cols else '',
                    'city': ns(ws.cell(row=r, column=cols.get('城市', 0)).value) if '城市' in cols else '',
                    'province': ns(ws.cell(row=r, column=cols.get('省份', 0)).value) if '省份' in cols else '',
                    'is_985': ns(ws.cell(row=r, column=cols.get('985', 0)).value) == '是' if '985' in cols else False,
                    'is_211': ns(ws.cell(row=r, column=cols.get('211', 0)).value) == '是' if '211' in cols else False,
                    'is_double_first': ns(ws.cell(row=r, column=cols.get('双一流', 0)).value) == '双一流' if '双一流' in cols else False,
                }
            wb.close()

        # Source 2: 2025 major data (has real 院校代码)
        f25 = file_by_kw('25年全国')
        if f25:
            print(f"  Reading 2025 major data for codes...")
            wb = openpyxl.load_workbook(f25, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hr, cols = find_header(ws)
            name_col = cols.get('院校名称', 0)
            code_col = cols.get('院校代码', 0)
            prov_col = cols.get('所在省', 0)
            attr_col = cols.get('公私性质', 0)

            if name_col and code_col:
                for r in range(hr + 1, ws.max_row + 1):
                    name = ns(ws.cell(row=r, column=name_col).value)
                    code = ns(ws.cell(row=r, column=code_col).value)
                    if not name or not code: continue
                    prov = ns(ws.cell(row=r, column=prov_col).value) if prov_col else ''
                    attr = ns(ws.cell(row=r, column=attr_col).value) if attr_col else ''
                    # Update existing or create new
                    existing = None
                    for c, s in schools_by_code.items():
                        if s['name'] == name:
                            existing = s
                            break
                    if existing:
                        if not existing['code'] or len(code) < len(existing['code']):
                            existing['code'] = code
                        if prov: existing['province'] = prov
                        if attr: existing['attribute'] = attr
                    else:
                        schools_by_code[code] = {
                            'code': code, 'name': name,
                            'type': '', 'attribute': attr, 'level': '',
                            'city': '', 'province': prov,
                            'is_985': False, 'is_211': False, 'is_double_first': False,
                        }
            wb.close()

        # Source 3: 2023 school-level file (code mapping)
        f23 = file_by_kw('2023-河北-院校')
        if f23:
            print(f"  Adding codes from 2023 school file...")
            wb = openpyxl.load_workbook(f23, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hr, cols = find_header(ws)
            name_col = cols.get('院校名称', 0)
            code_col = cols.get('院校代码', 0)
            if name_col and code_col:
                for r in range(hr + 1, ws.max_row + 1):
                    name = ns(ws.cell(row=r, column=name_col).value)
                    code = ns(ws.cell(row=r, column=code_col).value)
                    if name and code:
                        for c, s in schools_by_code.items():
                            if s['name'] == name and not s['code']:
                                s['code'] = code
                                break
            wb.close()

        # Insert schools
        school_list = list(schools_by_code.values())
        db.session.bulk_save_objects([School(**s) for s in school_list])
        db.session.commit()
        print(f"  -> {len(school_list)} schools inserted")

        # Build name->code map for admission data
        name_to_code = {s['name']: s['code'] for s in school_list}
        for s in school_list:
            if s['name'] not in name_to_code:
                name_to_code[s['name']] = s['code']

        # ---- ADMISSION RECORDS ----
        all_adm = []
        seen = set()

        def add_records(recs, label):
            added = 0
            for r in recs:
                key = (r['year'], r['subject'], r['min_score'], r['school_name'][:30], r['major_name'][:30])
                if key in seen: continue
                seen.add(key)
                code = name_to_code.get(r['school_name'], '')
                if not code:
                    code = f'X{abs(hash(r["school_name"])) % 9000 + 1000:04d}'
                r['school_code'] = code
                all_adm.append(r)
                added += 1
            print(f"  {label}: {added} records (total {len(all_adm)})")

        # 2023 major data
        f = file_by_kw('2023河北专业')
        if f:
            print(f"\n[2023 MAJOR] Loading...")
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hr, cols = find_header(ws)
            records = []
            for r in range(hr + 1, ws.max_row + 1):
                score = ni(ws.cell(row=r, column=cols.get('最低分', 0)).value)
                if not score or score < 100: continue
                name = ns(ws.cell(row=r, column=cols.get('院校名称', 0)).value)
                subj = ns(ws.cell(row=r, column=cols.get('科类', 0)).value)
                records.append({
                    'year': ni(ws.cell(row=r, column=cols.get('年份', 0)).value) or 2023,
                    'subject': '物理类' if '物理' in subj else '历史类',
                    'batch': ns(ws.cell(row=r, column=cols.get('批次', 0)).value),
                    'school_name': name,
                    'major_name': ns(ws.cell(row=r, column=cols.get('专业名称', 0)).value),
                    'major_code': '',
                    'min_score': score,
                    'min_rank': ni(ws.cell(row=r, column=cols.get('最低位次', 0)).value),
                    'subject_req': ns(ws.cell(row=r, column=cols.get('专业备注', 0)).value)[:50],
                })
            wb.close()
            add_records(records, '2023 major')

        # 2024 major data (undergrad + vocational)
        for track_label, track_kw in [('物理类', '物理'), ('历史类', '历史')]:
            for fname in os.listdir(DSK):
                if '专业录取数据' not in fname or '2024' not in fname: continue
                if track_kw not in fname: continue
                print(f"\n[2024 {track_label}] Loading {fname}...")
                path = os.path.join(DSK, fname)
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                hr, cols = find_header(ws)

                batch = '本科批' if '本科' in fname else '专科批'
                records = []
                for r in range(hr + 1, ws.max_row + 1):
                    sc = ni(ws.cell(row=r, column=cols.get('最低分', 0)).value)
                    if not sc or sc < 100: continue
                    records.append({
                        'year': 2024, 'subject': track_label, 'batch': batch,
                        'school_name': ns(ws.cell(row=r, column=cols.get('学校', 0)).value),
                        'major_name': ns(ws.cell(row=r, column=cols.get('专业', 0)).value),
                        'major_code': '',
                        'min_score': sc,
                        'min_rank': ni(ws.cell(row=r, column=cols.get('最低位次', 0)).value),
                        'subject_req': ns(ws.cell(row=r, column=cols.get('选科', 0)).value)[:50],
                    })
                wb.close()
                add_records(records, f'2024 {track_label}')

        # 2025 major data (the 48K record file)
        f25 = file_by_kw('25年全国')
        if f25:
            print(f"\n[2025 MAJOR] Loading...")
            wb = openpyxl.load_workbook(f25, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hr, cols = find_header(ws)
            records = []
            for r in range(hr + 1, ws.max_row + 1):
                sc = ni(ws.cell(row=r, column=cols.get('最低分', 0)).value)
                if not sc or sc < 100: continue
                subj_raw = ns(ws.cell(row=r, column=cols.get('科类', 0)).value)
                batch_raw = ns(ws.cell(row=r, column=cols.get('层次', 0)).value)
                if '本科' in batch_raw: batch = '本科批'
                elif '专科' in batch_raw or '高职' in batch_raw: batch = '专科批'
                else: batch = batch_raw
                records.append({
                    'year': 2025,
                    'subject': '物理类' if '物理' in subj_raw else '历史类',
                    'batch': batch,
                    'school_name': ns(ws.cell(row=r, column=cols.get('院校名称', 0)).value),
                    'major_name': ns(ws.cell(row=r, column=cols.get('专业', 0)).value),
                    'major_code': ns(ws.cell(row=r, column=cols.get('专业代码', 0)).value),
                    'min_score': sc,
                    'min_rank': ni(ws.cell(row=r, column=cols.get('最低位次', 0)).value),
                    'subject_req': ns(ws.cell(row=r, column=cols.get('选科要求', 0)).value)[:50],
                })
            wb.close()
            add_records(records, '2025 major')

        # 2025 vocational (official, D: drive)
        for track_cn, track_subj in [('物理', '物理类'), ('历史', '历史类')]:
            for fname in os.listdir(DDRV):
                if '2025' not in fname or track_cn not in fname: continue
                if not fname.endswith('.xlsx'): continue
                print(f"\n[2025 {track_subj} VOC] Loading {fname}...")
                path = os.path.join(DDRV, fname)
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                hr, cols = find_header(ws)
                records = []
                for r in range(hr + 1, ws.max_row + 1):
                    sc = ni(ws.cell(row=r, column=cols.get('投档最低分', 0)).value)
                    if not sc or sc < 100: continue
                    records.append({
                        'year': 2025, 'subject': track_subj, 'batch': '专科批',
                        'school_name': ns(ws.cell(row=r, column=cols.get('院校名称', 0)).value),
                        'major_name': ns(ws.cell(row=r, column=cols.get('专业名称', 0)).value),
                        'major_code': ns(ws.cell(row=r, column=cols.get('专业代码', 0)).value),
                        'min_score': sc,
                        'min_rank': None,
                        'subject_req': '',
                    })
                wb.close()
                add_records(records, f'2025 {track_subj} vocational')

        # Insert admission records
        print(f"\n[INSERT] {len(all_adm)} admission records into database...")
        batch = 2000
        for i in range(0, len(all_adm), batch):
            chunk = all_adm[i:i + batch]
            # Fill missing school_code
            for r in chunk:
                if 'school_code' not in r or not r['school_code']:
                    r['school_code'] = name_to_code.get(r['school_name'], f"X{abs(hash(r['school_name'])) % 9000 + 1000:04d}")
            db.session.bulk_insert_mappings(AdmissionRecord, chunk)
            if i % 10000 == 0:
                db.session.commit()
                print(f"  {i}/{len(all_adm)}")
        db.session.commit()

        # ---- SCORE DISTRIBUTION ----
        print(f"\n[SCORES] Loading score distribution...")
        # 2024 & 2025 from our v2 file
        path_sd = os.path.join(DSK, [f for f in os.listdir(DSK) if '一分一段' in f and f.endswith('.xlsx')][0])
        wb = openpyxl.load_workbook(path_sd, read_only=True, data_only=True)
        sd_records = []
        for sname in wb.sheetnames:
            if '对比' in sname or 'vs' in sname.lower(): continue
            year = 2024 if '2024' in sname else 2025
            subject = '物理类' if '物理' in sname else '历史类'
            ws = wb[sname]
            for r in range(4, ws.max_row + 1):
                score = ni(ws.cell(row=r, column=1).value)
                cnt = ni(ws.cell(row=r, column=2).value)
                cum = ni(ws.cell(row=r, column=3).value)
                if score and cum and 100 <= score <= 750:
                    sd_records.append({'year': year, 'subject': subject, 'score': score, 'count': cnt or 0, 'cumulative': cum})
        wb.close()
        # 2023 from D: drive
        path_2023 = os.path.join(DDRV, '2023-2025河北高考投档线大全.xlsx')
        if os.path.exists(path_2023):
            wb = openpyxl.load_workbook(path_2023, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                if '2023一分一段' not in sname: continue
                ws = wb[sname]
                subject = '物理类' if '物理' in sname else '历史类'
                for r in range(4, ws.max_row + 1):
                    score = ni(ws.cell(row=r, column=1).value)
                    cnt = ni(ws.cell(row=r, column=3).value)
                    cum = ni(ws.cell(row=r, column=4).value)
                    if score and cum:
                        sd_records.append({'year': 2023, 'subject': subject, 'score': score, 'count': cnt or 0, 'cumulative': cum})
            wb.close()
        db.session.bulk_insert_mappings(ScoreDistribution, sd_records)
        db.session.commit()
        print(f"  -> {len(sd_records)} score rows inserted")

        # ---- ENROLLMENT PLANS ----
        print(f"\n[PLANS] Loading enrollment plans...")
        ep_records = []

        # 2024 plans
        f = file_by_kw('河北省2024年高考招生计划')
        if f:
            print(f"  2024 plans...")
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                if '总表' in sname: continue
                ws = wb[sname]
                hr, cols = find_header(ws)
                for r in range(hr + 1, ws.max_row + 1):
                    sn = ns(ws.cell(row=r, column=cols.get('院校名称', 0)).value)
                    mn = ns(ws.cell(row=r, column=cols.get('专业名称', 0)).value)
                    if not sn and not mn: continue
                    ep_records.append({
                        'year': 2024, 'batch': sname, 'subject': '', 'plan_type': '',
                        'school_code': ns(ws.cell(row=r, column=cols.get('院校代码', 0)).value),
                        'school_name': sn,
                        'major_code': ns(ws.cell(row=r, column=cols.get('专业代码', 0)).value),
                        'major_name': mn,
                        'plan_count': ni(ws.cell(row=r, column=cols.get('计划数', 0)).value) or 0,
                        'subject_req': ns(ws.cell(row=r, column=cols.get('选科科目', 0)).value)[:50],
                        'tuition': ni(ws.cell(row=r, column=cols.get('学费', 0)).value),
                        'duration': ni(ws.cell(row=r, column=cols.get('学制', 0)).value),
                    })
            wb.close()

        # 2025 plans
        f = file_by_kw('2025年河北省招生计划')
        if f:
            print(f"  2025 plans...")
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hr, cols = find_header(ws)
            for r in range(hr + 1, ws.max_row + 1):
                sn = ns(ws.cell(row=r, column=cols.get('院校名称', 0)).value)
                mn = ns(ws.cell(row=r, column=cols.get('专业名称', 0)).value)
                if not sn and not mn: continue
                ep_records.append({
                    'year': 2025,
                    'batch': ns(ws.cell(row=r, column=cols.get('批次', 0)).value),
                    'subject': ns(ws.cell(row=r, column=cols.get('科目', 0)).value),
                    'plan_type': '',
                    'school_code': ns(ws.cell(row=r, column=cols.get('院校代码', 0)).value),
                    'school_name': sn,
                    'major_code': ns(ws.cell(row=r, column=cols.get('专业代码', 0)).value),
                    'major_name': mn,
                    'plan_count': ni(ws.cell(row=r, column=cols.get('计划数', 0)).value) or 0,
                    'subject_req': ns(ws.cell(row=r, column=cols.get('选科', 0)).value)[:50],
                    'tuition': ni(ws.cell(row=r, column=cols.get('学费', 0)).value),
                    'duration': ni(ws.cell(row=r, column=cols.get('学制', 0)).value),
                })
            wb.close()

        # 2023 plans
        f = file_by_kw('河北-2023-招生计划')
        if f:
            print(f"  2023 plans...")
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hr, cols = find_header(ws)
            for r in range(hr + 1, ws.max_row + 1):
                sn = ns(ws.cell(row=r, column=cols.get('院校名称', 0)).value)
                mn = ns(ws.cell(row=r, column=cols.get('专业名称', 0)).value)
                if not sn and not mn: continue
                ep_records.append({
                    'year': 2023,
                    'batch': ns(ws.cell(row=r, column=cols.get('录取批次', 0)).value),
                    'subject': ns(ws.cell(row=r, column=cols.get('科类名称', 0)).value),
                    'plan_type': '',
                    'school_code': ns(ws.cell(row=r, column=cols.get('院校代码', 0)).value),
                    'school_name': sn,
                    'major_code': ns(ws.cell(row=r, column=cols.get('专业代码', 0)).value),
                    'major_name': mn,
                    'plan_count': ni(ws.cell(row=r, column=cols.get('计划数', 0)).value) or 0,
                    'subject_req': ns(ws.cell(row=r, column=cols.get('选科科目', 0)).value)[:50],
                    'tuition': ni(ws.cell(row=r, column=cols.get('学费', 0)).value),
                    'duration': ni(ws.cell(row=r, column=cols.get('学制', 0)).value),
                })
            wb.close()

        for i in range(0, len(ep_records), 2000):
            db.session.bulk_insert_mappings(EnrollmentPlan, ep_records[i:i + 2000])
            if i % 20000 == 0: db.session.commit()
        db.session.commit()
        print(f"  -> {len(ep_records)} enrollment plan rows")

        # ---- FINAL SUMMARY ----
        print("\n" + "=" * 60)
        print("IMPORT COMPLETE")
        print(f"  Schools:             {School.query.count()}")
        print(f"  Admission Records:   {AdmissionRecord.query.count()}")
        print(f"  Score Distribution:  {ScoreDistribution.query.count()}")
        print(f"  Enrollment Plans:    {EnrollmentPlan.query.count()}")

        # Quick stats
        for y in [2023, 2024, 2025]:
            for s in ['物理类', '历史类']:
                c = AdmissionRecord.query.filter_by(year=y, subject=s).count()
                if c: print(f"    {y} {s}: {c} records")
        print("=" * 60)


if __name__ == '__main__':
    main()
