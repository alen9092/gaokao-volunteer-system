import json
from datetime import datetime
from io import BytesIO
from flask import Blueprint, request, send_file, flash, redirect, url_for
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from decorators import login_required

export_bp = Blueprint('export', __name__)


@export_bp.route('/export/excel', methods=['POST'])
@login_required
def export_excel():
    data = request.form.get('data', '[]')
    try:
        items = json.loads(data)
    except json.JSONDecodeError:
        flash('数据格式错误', 'danger')
        return redirect(url_for('simulation.simulate'))

    if not items:
        flash('请至少选择一项', 'warning')
        return redirect(url_for('simulation.simulate'))

    wb = Workbook()
    ws = wb.active
    ws.title = '志愿导出'

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    body_font = Font(name='微软雅黑', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['序号', '院校代码', '院校名称', '专业代码', '专业名称', '最低分', '最低位次', '批次']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for i, item in enumerate(items, 1):
        row = [i, item.get('school_code', ''), item.get('school_name', ''),
               item.get('major_code', ''), item.get('major_name', ''),
               item.get('min_score', ''), item.get('min_rank', ''), item.get('batch', '')]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col in (1, 2, 4, 6, 7) else 'left')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'志愿导出_{current_user.username}_{ts}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
